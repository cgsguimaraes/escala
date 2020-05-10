from openpyxl import Workbook
from django.http import HttpResponse
from app_escala.models import Especialidade, Escala, Prestador, Servico, Empresa


def plan_especialidade():
    especialidade = Especialidade.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Especialidade'

    ws['A1'] = "Id"
    ws['B1'] = "Nome"

    linha = 2
    
    for dados in especialidade:
        ws.cell(row=linha, column=1).value = dados.id
        ws.cell(row=linha, column=2).value = dados.nome
        linha +=1
    #ws = wb.create_sheet('Especialidades_Cadastradas')
    

    file_name = "Especialidade.xlsx"
    response = HttpResponse(content_type = "application/ms-excel")
    content = "attachment; filename = {0}".format(file_name)
    response['Content-Disposition'] = content
    wb.save(response)
    return response


def plan_prestador():
    prestador = Prestador.objects.all()
    prestador_qt = prestador.count()

    wb = Workbook()
    ws = wb.active
    ws.title = "Prestadores_Cadastrados"

    ws['A1'] = "Controle de Escala"
    ws['A2'] = 'Cadastro de Prestadores'
    ws['A3'] = "Quantidade Cadastrada: {}".format(prestador_qt)
    
    linha_prestador = 6

    ws['A{}'.format(linha_prestador-1)] = "Codigo"
    ws['B{}'.format(linha_prestador-1)] = "Nome"
    
    
    for prestador in prestador:
        ws.cell(row=linha_prestador, column=1).value = prestador.codigo
        ws.cell(row=linha_prestador, column=2).value = prestador.nome
        linha_prestador += 1

    file_name = "Prestadores.xlsx"
    response = HttpResponse(content_type = "application/ms-excel")
    content = "attachment; filename = {0}".format(file_name)
    response['Content-Disposition'] = content
    wb.save(response)
    return response
    

def plan_empresa():
    empresa = Empresa.objects.all()
    empresa_qt = empresa.count()
    wb = Workbook()
    ws = wb.active
    ws.title = "Empresa"

    ws['A1'] = "Controle de Escala"
    ws['A2'] = 'Cadastro das Empresas'
    ws['A3'] = "Quantidade Cadastrada: {}".format(empresa_qt)
    
    linha_empresa = 6

    ws['A{}'.format(linha_empresa - 1)] = "Codigo"
    ws['B{}'.format(linha_empresa - 1)] = "Nome"    

    for empresa in empresa:
        ws.cell(row=linha_empresa, column=1).value = empresa.codigo
        ws.cell(row=linha_empresa, column=2).value = empresa.nome
        linha_empresa +=1
    
    file_name = "Empresa.xlsx"
    response = HttpResponse(content_type='Application/ms-excel')
    content = "attachment; filename = {0}".format(file_name)
    response['Content-Disposition'] = content
    wb.save(response)
    return response
     

def plan_servico():
    servico = Servico.objects.all()
    servico_qt = servico.count()

    wb = Workbook()
    ws = wb.active
    ws.title = "Servico"

    ws['A1'] = "Controle de Escala"
    ws['A2'] = 'Cadastro dos Serviços'
    ws['A3'] = "Quantidade Cadastrada: {}".format(servico_qt)
    
    linha_servico = 6

    ws['A{}'.format(linha_servico - 1)] = "Codigo"
    ws['B{}'.format(linha_servico - 1)] = "Nome"

    for servico in servico:
        ws.cell(row=linha_servico, column=1).value = servico.id
        ws.cell(row=linha_servico, column=2).value = servico.nome
        linha_servico +=1

    file_name = "Servico.xlsx"
    response = HttpResponse(content_type='Application/ms-excel')
    content = "attatchment; filename = {0}".format(file_name)
    response['Content-Disposition'] = content
    wb.save(response)
    return response




def plan_escala():

    especialidade = Especialidade.objects.all()
    prestador = Prestador.objects.all()
    servico = Servico.objects.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Especialidades_Cadastradas'
    
    #.merge_cells('A1:E1')

    ws['A1'] = 'ID'
    ws['B1'] = 'ESPECIALIDADE'

    cont = 2
    
    for especialidade in especialidade:
        ws.cell(row=cont, column=1).value = especialidade.id
        ws.cell(row=cont, column=2).value = especialidade.nome

        cont+=1

    sheet_escala = wb.create_sheet('Modelo', 0)
    sheet_escala['A1'] = 'MODELO DE IMPORTAÇÃO DE ESCALA'
    sheet_escala['A3'] = 'criacao'
    sheet_escala['B3'] = 'modificacao'
    sheet_escala['c3'] = 'situacao'
    sheet_escala['c4'] = 'True'
    sheet_escala['d3'] = 'data_producao'
    sheet_escala['d4'] = '30/04/2025'
    sheet_escala['e3'] = 'data_pagamento'
    sheet_escala['e4'] = '30/04/2025'
    sheet_escala['f3'] = 'quantidade'
    sheet_escala['f4'] = '12'
    sheet_escala['g3'] = 'valor'
    sheet_escala['g4'] = '1400,00'
    sheet_escala['h3'] = 'porta'
    sheet_escala['h4'] = 'AMBULATORIO'
    sheet_escala['i3'] = 'convenio'
    sheet_escala['i4'] = 'SUS'

    sheet_escala['j3'] = 'escala'
    sheet_escala['j4'] = 'HORAS'

    sheet_escala['k3'] = 'especialidade_id'
    sheet_escala['k4'] = '1'
    sheet_escala['l3'] = 'prestador_id'
    sheet_escala['l4'] = '15802'
    sheet_escala['m3'] = 'procedimentosus'
    sheet_escala['n3'] = 'empresa_multi_id'
    sheet_escala['n4'] = '1'
    sheet_escala['o3'] = 'servico_escala_id'
    sheet_escala['o4'] = '1'
    sheet_escala['p3'] = 'Aplicar Fórmula'
    sheet_escala['p4'] = '= A4 & "|" & B4 & "|" & C4 & "|" & TEXTO(D4,"AAAA-MM-DD") & "|" & TEXTO(E4, "AAAA-MM-DD") & "|" & F4 & "|" & SUBSTITUTE(G4,",",".") & "|" & H4 & "|" & I4 & "|" & J4 & "|" & K4 & "|" & L4 & "|" & M4 & "|" & N4 & "|" & O4'
                    
    ws_prestador = wb.create_sheet('Prestadores_Cadastrador', 2)
    ws_prestador['A1'] = "CODIGO"
    ws_prestador['B1'] = "PRESTADOR"

    linha = 2
    for prestador in prestador:
        ws_prestador.cell(row=linha, column=1).value = prestador.codigo
        ws_prestador.cell(row=linha, column=2).value = prestador.nome
        linha +=1 
    
    ws_servico = wb.create_sheet('Servicos_Cadastrados', 3)
    ws_servico['A1'] = "ID"
    ws_servico['B1'] = "DESCRIÇÃO"

    linha_servico = 2
    for servico in servico:
        ws_servico.cell(row=linha_servico, column=1).value = servico.id
        ws_servico.cell(row=linha_servico, column=2).value = servico.nome
        linha_servico += 1


    file_name = "Modelo_Escala.xlsx"
    response = HttpResponse(content_type = "application/ms-excel")
    content = "attachment; filename = {0}".format(file_name)
    response['Content-Disposition'] = content
    wb.save(response)
    return response
